#!/usr/bin/env python3
"""
Банковский Радар — агент автоматического обновления.
Стратегия: ищет PDF на e-disclosure.ru и IR-страницах банков,
извлекает метрики через pdfplumber, обновляет data.json.
Показывает данные на две даты: базовую (годовая) и последнюю доступную.
"""

import json, re, sys, time, hashlib, io
import urllib.request
import urllib.parse, urllib.error
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False
    print("WARN: pdfplumber не установлен")

MSK    = timezone(timedelta(hours=3))
TODAY  = datetime.now(MSK).strftime("%d.%m.%Y")
NOW    = datetime.now(MSK).strftime("%d.%m.%Y %H:%M МСК")
ROOT   = Path(__file__).parent.parent
OUTPUT = ROOT / "data.json"
CACHE  = ROOT / ".pdf_cache.json"

UA = ("Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

# ── РЕЕСТР БАНКОВ ────────────────────────────────────────────────────
# e_disclosure_id — ID на e-disclosure.ru для поиска PDF
# Базовые данные МСФО 2025 (годовые) — эталон
BANKS = [
  {"r":1,  "n":"Сбербанк",       "lo":"С","lc":"#21A038","tp":"state",  "pub":1,"pf":0,
   "e_id":"1481", "ticker":"SBER",
   "ir_urls":["https://www.sberbank.ru/ru/investor-relations","https://smart-lab.ru/q/SBER/f/y/"],
   "base":{"roe":22.7,"nim":6.2,"cir":30.3,"cor":1.3,"h20":14.6,"src":"МСФО фев.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":0},"st":"data"},

  {"r":2,  "n":"ВТБ",            "lo":"В","lc":"#003087","tp":"state",  "pub":1,"pf":0,
   "e_id":"1000", "ticker":"VTBR",
   "ir_urls":["https://www.vtb.ru/ir/statements/results/","https://smart-lab.ru/q/VTBR/f/y/"],
   "base":{"roe":18.3,"nim":1.4,"cir":47.3,"cor":1.1,"h20":9.8,"src":"МСФО фев.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":0},"st":"data"},

  {"r":3,  "n":"Газпромбанк",    "lo":"Г","lc":"#1B4D9A","tp":"state",  "pub":0,"pf":0,
   "e_id":"2547", "ticker":None,
   "ir_urls":["https://www.gazprombank.ru/ir/results/","https://bosfera.ru/bo/gazprombank"],
   "base":{"roe":10.1,"nim":3.0,"cir":48.6,"cor":None,"h20":11.5,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":1,"he":1},"st":"part"},

  {"r":4,  "n":"Альфа-Банк",     "lo":"А","lc":"#EF3124","tp":"private","pub":0,"pf":0,
   "e_id":None, "ticker":None,
   "ir_urls":["https://alfabank.ru/alfa-investor/","https://bosfera.ru/bo/alfa-bank"],
   "base":{"roe":21.0,"nim":5.8,"cir":42.0,"cor":2.7,"h20":13.0,"src":"МСФО апр.2026","period":"2025"},
   "est":{"re":1,"ne":1,"ce":1,"oe":1,"he":1},"st":"part"},

  {"r":5,  "n":"Россельхозбанк", "lo":"Р","lc":"#2E7D32","tp":"state",  "pub":0,"pf":0,
   "e_id":"2749", "ticker":None,
   "ir_urls":["https://www.rshb.ru/about/reports-conclusion/msfo/","https://bosfera.ru/bo/rosselkhozbank"],
   "base":{"roe":14.5,"nim":2.7,"cir":50.0,"cor":None,"h20":12.4,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":1,"he":1},"st":"part"},

  {"r":6,  "n":"МКБ",            "lo":"М","lc":"#7B1FA2","tp":"private","pub":1,"pf":0,
   "e_id":"2096", "ticker":"CBOM",
   "ir_urls":["https://ir.mkb.ru/investor-relations/reports/ifrs","https://smart-lab.ru/q/CBOM/f/y/"],
   "base":{"roe":8.9,"nim":1.9,"cir":33.8,"cor":15.1,"h20":11.0,"src":"Smart-lab LTM 2026","period":"LTM"},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":0},"st":"data"},

  {"r":7,  "n":"Т-Технологии",   "lo":"Т","lc":"#111111","tp":"private","pub":1,"pf":0,
   "e_id":None, "ticker":"T",
   "ir_urls":["https://t-technologies.ru/results/","https://smart-lab.ru/q/T/f/y/"],
   "base":{"roe":29.1,"nim":10.8,"cir":34.7,"cor":6.5,"h20":12.5,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":1},"st":"data"},

  {"r":8,  "n":"Совкомбанк",     "lo":"С","lc":"#E65100","tp":"private","pub":1,"pf":0,
   "e_id":None, "ticker":"SVCB",
   "ir_urls":["https://sovcombank.ru/about/disclosure/financial-statements/","https://smart-lab.ru/q/SVCB/f/y/"],
   "base":{"roe":15.0,"nim":5.3,"cir":57.0,"cor":2.5,"h20":13.0,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":1},"st":"data"},

  {"r":9,  "n":"Банк ДОМ.РФ",   "lo":"Д","lc":"#1565C0","tp":"state",  "pub":1,"pf":1,
   "e_id":None, "ticker":None,
   "ir_urls":["https://domrfbank.ru/about/information/msfo/","https://domrfbank.ru/about/press/"],
   "base":{"roe":21.6,"nim":3.7,"cir":28.3,"cor":0.7,"h20":13.5,"src":"МСФО фев.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":1},"st":"data"},

  {"r":10, "n":"Промсвязьбанк",            "lo":"П","lc":"#AD1457","tp":"state",  "pub":0,"pf":0,
   "e_id":None, "ticker":None,
   "ir_urls":["https://www.psbank.ru/Bank/Investors/IFRS","https://www.psbank.ru/bank/investors/ras"],
   "base":{"roe":30.8,"nim":3.45,"cir":45.1,"cor":None,"h20":None,"src":"МСФО 1кв.2026 (устойч.)","period":"1кв2026"},
   "est":{"re":1,"ne":1,"ce":1,"oe":1,"he":0},"st":"part"},

  {"r":11, "n":"Юникредит Банк", "lo":"U","lc":"#c00000","tp":"foreign","pub":0,"pf":0,
   "e_id":None, "ticker":None,
   "ir_urls":["https://www.unicreditbank.ru/ru/about/shareholders-and-investors.html"],
   "base":{"roe":None,"nim":None,"cir":None,"cor":None,"h20":None,"src":"уход до сер.2026","period":None},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":0},"st":"exit"},

  {"r":12, "n":"Райффайзенбанк", "lo":"R","lc":"#FFD700","tp":"foreign","pub":0,"pf":0,
   "e_id":"3292", "ticker":None,
   "ir_urls":["https://www.raiffeisen.ru/about/investors/information/","https://bosfera.ru/bo/raiffeisenbank"],
   "base":{"roe":12.9,"nim":None,"cir":None,"cor":None,"h20":65.6,"src":"РСБУ 1кв.2026","period":"1кв2026"},
   "est":{"re":1,"ne":0,"ce":0,"oe":0,"he":0},"st":"part"},

  {"r":13, "n":"БСПБ",           "lo":"Б","lc":"#0D47A1","tp":"private","pub":1,"pf":0,
   "e_id":None, "ticker":"BSPB",
   "ir_urls":["https://www.bspb.ru/investors/financial-statements/IFRS","https://smart-lab.ru/q/BSPB/f/y/"],
   "base":{"roe":18.1,"nim":7.4,"cir":29.0,"cor":2.0,"h20":14.0,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":0,"ne":1,"ce":1,"oe":1,"he":1},"st":"data"},

  {"r":14, "n":"МТС Банк",       "lo":"М","lc":"#E53935","tp":"private","pub":1,"pf":0,
   "e_id":None, "ticker":"MBNK",
   "ir_urls":["https://www.mtsbank.ru/investors-and-shareholders/results/","https://smart-lab.ru/q/MBNK/f/y/"],
   "base":{"roe":14.5,"nim":7.2,"cir":35.7,"cor":5.5,"h20":12.0,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":0,"ne":0,"ce":0,"oe":1,"he":1},"st":"data"},

  {"r":15, "n":"Новикомбанк",    "lo":"Н","lc":"#0D47A1","tp":"state",  "pub":0,"pf":0,
   "e_id":"9789", "ticker":None,
   "ir_urls":["https://novikombank.ru/about/disclosure/financial-reports/","https://ratings.ru/ratings/press-releases/?search=новикомбанк"],
   "base":{"roe":25.0,"nim":None,"cir":None,"cor":None,"h20":15.2,"src":"НКР 9м.2025","period":"9м2025"},
   "est":{"re":1,"ne":0,"ce":0,"oe":1,"he":0},"st":"part"},

  {"r":16, "n":"Уралсиб",        "lo":"У","lc":"#00695C","tp":"private","pub":0,"pf":0,
   "e_id":None, "ticker":None,
   "ir_urls":["https://uralsib.ru/about/investors/","https://bosfera.ru/bo/bank-uralsib"],
   "base":{"roe":10.1,"nim":None,"cir":None,"cor":None,"h20":None,"src":"РСБУ апр.2026","period":"2025"},
   "est":{"re":1,"ne":0,"ce":0,"oe":0,"he":0},"st":"part"},

  {"r":17, "n":"Ак Барс",        "lo":"А","lc":"#558B2F","tp":"private","pub":0,"pf":0,
   "e_id":"2975", "ticker":None,
   "ir_urls":["https://www.akbars.ru/about/free-info/reports/","https://cbonds.ru/company/AKBARS/"],
   "base":{"roe":14.3,"nim":None,"cir":None,"cor":None,"h20":13.9,"src":"МСФО мар.2026","period":"2025"},
   "est":{"re":1,"ne":0,"ce":0,"oe":1,"he":0},"st":"part"},

  {"r":18, "n":"БМ-Банк",        "lo":"О","lc":"#4527A0","tp":"state",  "pub":0,"pf":0,
   "e_id":None, "ticker":None,
   "ir_urls":[],
   "base":{"roe":None,"nim":None,"cir":None,"cor":None,"h20":None,"src":"правопреемник Открытия с 01.01.2025, присоединение к ВТБ отложено","period":None},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":0},"st":"nd"},

  {"r":19, "n":"Банк Россия",    "lo":"Р","lc":"#1A237E","tp":"private","pub":0,"pf":0,
   "e_id":None, "ticker":None,
   "ir_urls":[],
   "base":{"roe":None,"nim":None,"cir":None,"cor":None,"h20":None,"src":"данные закрыты (санкции с 2014)","period":None},
   "est":{"re":0,"ne":0,"ce":0,"oe":0,"he":0},"st":"nd"},
]


# ── SMART-LAB ТИКЕРЫ ─────────────────────────────────────────────────
# Биржевые банки с публичной МСФО-таблицей на Smart-lab
SMARTLAB_TICKERS = {
    1:  'SBER',   # Сбербанк
    2:  'VTBR',   # ВТБ
    6:  'CBOM',   # МКБ
    7:  'T',      # Т-Технологии
    8:  'SVCB',   # Совкомбанк
    13: 'BSPB',   # БСПБ
    14: 'MBNK',   # МТС Банк
}

def parse_smartlab(ticker: str) -> dict:
    """Парсит LTM-данные МСФО со страницы Smart-lab."""
    url = f'https://smart-lab.ru/q/{ticker}/f/y/'
    result = {}
    html = fetch_html(url)
    if not html:
        return result

    import re

    # ROE
    for pat in [r'ROE[^%\d]{0,50}([\d.]+)\s*%',
                r'>\s*([\d.]+)\s*%\s*</td>.*?ROE']:
        m = re.search(pat, html, re.S | re.I)
        if m:
            try:
                v = float(m.group(1))
                if 0.5 < v < 60: result['roe'] = round(v, 1); break
            except: pass

    # NIM — чистая процентная маржа
    for pat in [r'[Чч]ист\.?\s*проц\.?\s*маржа[^%\d]{0,60}([\d.]+)\s*%',
                r'NIM[^%\d]{0,30}([\d.]+)\s*%']:
        m = re.search(pat, html, re.S | re.I)
        if m:
            try:
                v = float(m.group(1))
                if 0.5 < v < 20: result['nim'] = round(v, 2); break
            except: pass

    # CIR
    for pat in [r'(?:CIR|[Рр]асходы.{0,10}[Дд]оходы)[^%\d]{0,40}([\d.]+)\s*%']:
        m = re.search(pat, html, re.S | re.I)
        if m:
            try:
                v = float(m.group(1))
                if 10 < v < 100: result['cir'] = round(v, 1); break
            except: pass

    # COR
    for pat in [r'(?:CoR|COR|[Сс]тоимость\s+риска)[^%\d]{0,40}([\d.]+)\s*%']:
        m = re.search(pat, html, re.S | re.I)
        if m:
            try:
                v = float(m.group(1))
                if 0 < v < 25: result['cor'] = round(v, 2); break
            except: pass

    # H20 / достаточность капитала
    for pat in [r'[Дд]ост\.?\s*осн\.?\s*капитала[^%\d]{0,40}([\d.]+)\s*%',
                r'H20\.0[^%\d]{0,20}([\d.]+)\s*%']:
        m = re.search(pat, html, re.S | re.I)
        if m:
            try:
                v = float(m.group(1))
                if 8 < v < 30: result['h20'] = round(v, 2); break
            except: pass

    # Если ROE не нашли — считаем из прибыли и капитала
    if 'roe' not in result:
        pm = re.search(r'[Чч]истая\s+прибыль[^<\d]{0,30}([\d\s]+\.?\d*)\s*(?:млрд)', html)
        cm = re.search(r'[Кк]апитал[^<\d]{0,20}([\d\s]+\.?\d*)\s*(?:млрд)', html)
        if pm and cm:
            try:
                p = float(pm.group(1).replace(' ', ''))
                c = float(cm.group(1).replace(' ', ''))
                if c > 0 and 0 < p < c:
                    result['roe'] = round(p / c * 100, 1)
            except: pass

    return result


# ── ПАТТЕРНЫ ИЗВЛЕЧЕНИЯ МЕТРИК ───────────────────────────────────────
PATTERNS = {
    "roe": [
        r'(?:ROE|рентабельност\w+\s+(?:собственного\s+)?капитал\w*)[^\d]{0,50}?([\d]+[,.][\d]+)\s*%',
        r'(?:return\s+on\s+equity)[^\d]{0,30}?([\d]+[,.][\d]+)\s*%',
    ],
    "nim": [
        r'(?:NIM|чист\w+\s+процентн\w+\s+марж\w+)[^\d]{0,50}?([\d]+[,.][\d]+)\s*%',
        r'(?:net\s+interest\s+margin)[^\d]{0,30}?([\d]+[,.][\d]+)\s*%',
    ],
    "cir": [
        r'(?:CIR|CTI|cost.to.income|расход\w+\s+к\s+доход\w+)[^\d]{0,50}?([\d]+[,.][\d]+)\s*%',
        r'(?:отношение\s+(?:операционных\s+)?расход\w+)[^\d]{0,50}?([\d]+[,.][\d]+)\s*%',
    ],
    "cor": [
        r'(?:COR|стоимост\w+\s+риска|cost\s+of\s+risk)[^\d]{0,40}?([\d]+[,.][\d]+)\s*%',
    ],
    "h20": [
        r'(?:H20\.0|Н20\.0|норматив\s+достаточности)[^\d]{0,40}?(1[\d][,.][\d]+)\s*%',
        r'(?:H1\.0|Н1\.0|достаточност\w+\s+капитал\w+)[^\d]{0,30}?(1[\d][,.][\d]+)\s*%',
    ],
}

RANGES = {"roe":(0,60),"nim":(0.3,25),"cir":(5,95),"cor":(0.1,20),"h20":(8,35)}

def fetch(url, timeout=20):
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": UA,
            "Accept-Language": "ru-RU,ru;q=0.9",
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
        })
        with urllib.request.urlopen(req, timeout=timeout) as r:
            raw = r.read()
            enc = r.headers.get_content_charset() or "utf-8"
            return raw.decode(enc, errors="replace")
    except Exception as e:
        print(f"    WARN {url[:65]}: {e}")
        return None

def fetch_binary(url, timeout=40):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": UA})
        with urllib.request.urlopen(req, timeout=timeout) as r:
            return r.read()
    except Exception as e:
        print(f"    WARN binary {url[:65]}: {e}")
        return None

def md5(data):
    return hashlib.md5(data).hexdigest()

def exnum(text, pats):
    for p in pats:
        for m in re.finditer(p, text, re.I | re.M):
            try:
                v = float(m.group(1).replace(",","."))
                return v
            except: pass
    return None

def validate(metric, val):
    lo, hi = RANGES.get(metric, (0,1000))
    return lo <= val <= hi

def extract_metrics(text):
    found = {}
    for m, pats in PATTERNS.items():
        v = exnum(text, pats)
        if v is not None and validate(m, v):
            found[m] = round(v, 2)
    return found

def pdf_to_text(pdf_bytes):
    if not PDF_OK: return ""
    parts = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:40]:
                t = page.extract_text()
                if t: parts.append(t)
    except Exception as e:
        print(f"    WARN pdfplumber: {e}")
    return "\n".join(parts)

def detect_period(text):
    """Определяет отчётный период из текста PDF."""
    # 1 квартал 2026
    if re.search(r'(?:1|первый|I)\s*(?:квартал|кв\.?)\s*2026|31\s*март\w*\s*2026|March\s*31,?\s*2026', text, re.I):
        return "1 кв.2026", "31.03.2026"
    # Полугодие 2025
    if re.search(r'(?:6\s*мес|полугоди\w+|II\s*кв)\s*2025|30\s*июн\w*\s*2025|June\s*30,?\s*2025', text, re.I):
        return "1П 2025", "30.06.2025"
    # 9 месяцев 2025
    if re.search(r'(?:9\s*мес|9\s*месяц\w+|III\s*кв)\s*2025|30\s*сент\w*\s*2025|September\s*30,?\s*2025', text, re.I):
        return "9М 2025", "30.09.2025"
    # Год 2025
    if re.search(r'(?:год|year|annual)\w*\s*2025|31\s*декабр\w*\s*2025|December\s*31,?\s*2025', text, re.I):
        return "2025", "31.12.2025"
    # 1 квартал 2025
    if re.search(r'(?:1|первый|I)\s*(?:квартал|кв\.?)\s*2025|31\s*март\w*\s*2025', text, re.I):
        return "1 кв.2025", "31.03.2025"
    return None, None

def find_pdfs_on_edisclosure(e_id, bank_name):
    """Ищет PDF МСФО отчётности на e-disclosure.ru."""
    if not e_id:
        return []
    url = f"https://e-disclosure.ru/portal/files.aspx?id={e_id}&type=4"
    html = fetch(url)
    if not html:
        return []
    # Ищем ссылки на PDF
    links = re.findall(r'href=["\']([^"\']*fileload\.ashx\?fileid=\d+)["\']', html, re.I)
    links += re.findall(r'href=["\']([^"\']*\.pdf[^"\']*)["\']', html, re.I)
    # Фильтруем по ключевым словам МСФО
    msfo_links = []
    for link in links[:20]:
        if not link.startswith("http"):
            link = "https://e-disclosure.ru" + link
        msfo_links.append(link)
    return list(dict.fromkeys(msfo_links))[:5]  # убираем дубли

def process_bank_pdf(bank):
    """Обрабатывает один банк: ищет PDF, извлекает метрики."""
    name = bank["n"]
    e_id = bank.get("e_id")
    print(f"  [{bank['r']}] {name}...")

    if not e_id:
        print(f"    → нет e_id, пропускаем")
        return None

    pdf_links = find_pdfs_on_edisclosure(e_id, name)
    print(f"    → найдено PDF: {len(pdf_links)}")

    cache = load_cache()
    best = None  # лучший найденный результат

    for pdf_url in pdf_links:
        cache_key = "v2_" + md5(pdf_url.encode())
        if cache_key in cache and cache[cache_key].get("metrics"):
            cached = cache[cache_key]
            print(f"    → из кэша: {cached.get('period','?')} — {cached.get('metrics',{})}")
            if best is None or _is_newer(cached.get("period_date"), best.get("period_date")):
                best = cached
            continue

        print(f"    → скачиваю {pdf_url[-60:]}...")
        pdf_bytes = fetch_binary(pdf_url)
        if not pdf_bytes or len(pdf_bytes) < 5000:
            continue

        text = pdf_to_text(pdf_bytes)
        if not text:
            cache[cache_key] = {"url": pdf_url, "metrics": {}, "period": None}
            save_cache(cache)
            continue

        metrics = extract_metrics(text)
        period_label, period_date = detect_period(text)

        result = {
            "url": pdf_url,
            "metrics": metrics,
            "period": period_label,
            "period_date": period_date,
            "bank": name,
        }
        cache[cache_key] = result
        save_cache(cache)

        if metrics:
            print(f"    → {period_label}: {metrics}")
            if best is None or _is_newer(period_date, best.get("period_date")):
                best = result

        time.sleep(1.5)

    return best

def _is_newer(date_str, than_str):
    """Сравнивает даты в формате ДД.ММ.ГГГГ."""
    if not date_str: return False
    if not than_str: return True
    try:
        d1 = datetime.strptime(date_str, "%d.%m.%Y")
        d2 = datetime.strptime(than_str, "%d.%m.%Y")
        return d1 > d2
    except:
        return False

def load_cache():
    if CACHE.exists():
        try: return json.loads(CACHE.read_text("utf-8"))
        except: pass
    return {}

def save_cache(cache):
    CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), "utf-8")

def compute_kpi(banks):
    roes  = [b["roe"] for b in banks if b.get("roe")]
    rorwa = [b for b in banks if b.get("rorwa")]
    pfcir = [b for b in banks if b.get("pf") and b.get("cir")]
    h20s  = [b for b in banks if b.get("h20")]
    med   = round(sorted(roes)[len(roes)//2], 1) if roes else None
    ldr   = max(rorwa, key=lambda b: b["rorwa"]) if rorwa else None
    bcir  = min(pfcir, key=lambda b: b["cir"]) if pfcir else None
    mnh   = min(h20s,  key=lambda b: b["h20"]) if h20s  else None
    return {
        "median_roe": med,
        "leader_rorwa_bank":  ldr["n"]    if ldr  else "—",
        "leader_rorwa_val":   ldr["rorwa"]if ldr  else None,
        "best_cir_pf_bank":   bcir["n"]  if bcir else "—",
        "best_cir_pf_val":    bcir["cir"]if bcir else None,
        "min_h20_bank":       mnh["n"]   if mnh  else "—",
        "min_h20_val":        mnh["h20"] if mnh  else None,
        "dom_rf_pf_portfolio": "1.9 трлн",
    }

def build_signals(banks):
    sigs = []
    wr = [b for b in banks if b.get("rorwa")]
    if wr:
        top = max(wr, key=lambda b: b["rorwa"])
        sigs.append({"bank":top["n"],"logo":top["lo"],"lc":top["lc"],
            "type":"metric","priority":"high","score":95,
            "text":f"Лидер iRoRWA: {top['rorwa']:.2f}% (ROE {top['roe']}% × H20.0 {top['h20']}%). Наилучшая эффективность капитала в ТОП-20.",
            "tags":["iRoRWA","лидер"],"time":TODAY})
    wh = [b for b in banks if b.get("h20")]
    if wh:
        mn = min(wh, key=lambda b: b["h20"])
        sigs.append({"bank":mn["n"],"logo":mn["lo"],"lc":mn["lc"],
            "type":"risk","priority":"high" if mn["h20"]<11 else "med","score":85,
            "text":f"Минимальный буфер капитала: H20.0={mn['h20']}% при минимуме 9.25%. Буфер {round(mn['h20']-9.25,2)} пп.",
            "tags":["H20","капитал","риск"],"time":TODAY})
    pf = [b for b in banks if b.get("pf") and b.get("cir")]
    if pf:
        bc = min(pf, key=lambda b: b["cir"])
        sigs.append({"bank":bc["n"],"logo":bc["lo"],"lc":bc["lc"],
            "type":"metric","priority":"high","score":90,
            "text":f"Лучший CIR в ПФ: {bc['cir']}%. Эскроу-модель структурно снижает COR={bc.get('cor','—')}%.",
            "tags":["CIR","ПФ","эффективность"],"time":TODAY})
    for b in banks:
        if b.get("st")=="loss":
            sigs.append({"bank":b["n"],"logo":b["lo"],"lc":b["lc"],
                "type":"risk","priority":"high","score":88,
                "text":f"Убыток/слабые результаты по МСФО: {b.get('src','')}.",
                "tags":["убыток","риск"],"time":TODAY})
    return sigs[:8]


# ── ПАРСИНГ ОТКРЫТЫХ ИСТОЧНИКОВ ДЛЯ АК БАРС / ЭКСПОБАНК / СИНАРА ────────
OPEN_SOURCES = {
    15: {
        "name": "Новикомбанк",
        "urls": ["https://novikombank.ru/about/disclosure/financial-reports/",
                 "https://cbonds.ru/company/novikombank/"],
        "news_url": "https://cbonds.ru/news/?search=Новикомбанк+МСФО",
    },
    17: {
        "name": "Ак Барс",
        "urls": ["https://www.akbars.ru/about/free-info/reports/",
                 "https://cbonds.ru/company/AKBARS/"],
        "news_url": "https://cbonds.ru/news/?search=Ак+Барс+МСФО",
    },
}

def parse_open_source_bank(rank: int) -> dict:
    """
    Парсит данные банков без биржевого листинга из открытых источников:
    официальные сайты, cbonds.ru, пресс-релизы.
    """
    src_cfg = OPEN_SOURCES.get(rank)
    if not src_cfg:
        return {}

    name = src_cfg["name"]
    result = {}

    for url in src_cfg["urls"]:
        print(f"    → {url[:60]}...")
        html = fetch(url)
        if not html:
            time.sleep(1)
            continue

        # Убираем HTML теги
        text = re.sub(r'<[^>]+>', ' ', html)
        text = re.sub(r'\s+', ' ', text)

        # Ищем метрики
        metrics = extract_metrics(text)
        if metrics:
            print(f"    ✓ найдено: {metrics}")
            result.update(metrics)

        # Ищем PDF ссылки для дальнейшего парсинга
        pdf_links = re.findall(
            r'href=["\']([^"\']+\.pdf[^"\']*)["\']|'
            r'href=["\']([^"\']+msfo[^"\']*)["\']|'
            r'href=["\']([^"\']+отчет[^"\']*)["\']i',
            html, re.I
        )
        base_domain = re.match(r'https?://[^/]+', url)
        base = base_domain.group(0) if base_domain else ""

        for link_groups in pdf_links[:5]:
            link = next((g for g in link_groups if g), None)
            if not link:
                continue
            if link.startswith("/"):
                link = base + link
            elif not link.startswith("http"):
                link = base + "/" + link

            if ".pdf" in link.lower():
                print(f"    → PDF: {link[-60:]}...")
                pdf_bytes = fetch_binary(link)
                if pdf_bytes and len(pdf_bytes) > 5000:
                    pdf_text = extract_pdf_text(pdf_bytes)
                    if pdf_text:
                        pdf_metrics = extract_metrics(pdf_text)
                        if pdf_metrics:
                            print(f"    ✓ PDF метрики: {pdf_metrics}")
                            result.update(pdf_metrics)
                time.sleep(1)

        if result:
            break
        time.sleep(1.5)

    # Парсим новостную ленту если ничего не нашли
    if not result and src_cfg.get("news_url"):
        html = fetch(src_cfg["news_url"])
        if html:
            text = re.sub(r'<[^>]+>', ' ', html)
            text = re.sub(r'\s+', ' ', text)
            metrics = extract_metrics(text)
            if metrics:
                print(f"    ✓ новости: {metrics}")
                result.update(metrics)

    return result



# ── ПАРСИНГ ДАННЫХ ПФ ИЗ ЦБ РФ ───────────────────────────────────────
# ЦБ РФ публикует Excel-файл с данными по банкам ежемесячно
# URL: cbr.ru/statistics/bank_sector/equity_const_financing/
# Файл содержит: банк, кредиты застройщикам (млрд руб.), счета эскроу

CBR_PF_URLS = [
    # Пробуем последние месяцы по шаблону
    "https://www.cbr.ru/vfs/statistics/bank_sector/pf/pf_{year}{month:02d}.xlsx",
    "https://cbr.ru/vfs/statistics/bank_sector/equity_const_financing/pf_{year}{month:02d}.xlsx",
]

# Статичные данные из последнего отчёта ЦБ РФ 1кв.2025
# Обновляются вручную при выходе нового квартального отчёта
# Источник: cbr.ru/collection/collection/file/55905/pf_2025_q1.pdf
CBR_PF_SHARES = {
    "Сбербанк":       {"pf_share": 40.0, "pf_trn": 3.48, "asset_rank": 1},
    "ВТБ":            {"pf_share": 14.0, "pf_trn": 1.22, "asset_rank": 2},
    "Банк ДОМ.РФ":   {"pf_share": 17.0, "pf_trn": 1.48, "asset_rank": 9},
    "Газпромбанк":    {"pf_share": 6.0,  "pf_trn": 0.52, "asset_rank": 3},
    "Альфа-Банк":     {"pf_share": 4.0,  "pf_trn": 0.35, "asset_rank": 4},
    "Промсвязьбанк":  {"pf_share": 3.0,  "pf_trn": 0.26, "asset_rank": 10},
    "Россельхозбанк": {"pf_share": 2.0,  "pf_trn": 0.18, "asset_rank": 5},
    "МКБ":            {"pf_share": 1.0,  "pf_trn": 0.09, "asset_rank": 6},
    "Совкомбанк":     {"pf_share": 1.0,  "pf_trn": 0.09, "asset_rank": 8},
    "БСПБ":           {"pf_share": 0.5,  "pf_trn": 0.04, "asset_rank": 13},
    "МТС Банк":       {"pf_share": 0.5,  "pf_trn": 0.04, "asset_rank": 14},
    "Ак Барс":        {"pf_share": 0.3,  "pf_trn": 0.03, "asset_rank": 17},
    "Новикомбанк":    {"pf_share": 0.2,  "pf_trn": 0.02, "asset_rank": 15},
    "Уралсиб":        {"pf_share": 0.2,  "pf_trn": 0.02, "asset_rank": 16},
}

def fetch_cbr_pf_excel() -> dict:
    """
    Пытается скачать Excel с сайта ЦБ РФ с актуальными данными по ПФ.
    Возвращает словарь {bank_name: {pf_share, pf_trn, asset_rank}} или пустой dict.
    """
    from datetime import datetime, timedelta
    import urllib.request, re

    # Пробуем последние 3 месяца
    today = datetime.today()
    for delta in [0, 1, 2]:
        dt = today - timedelta(days=30*delta)
        year, month = dt.year, dt.month

        for url_tpl in CBR_PF_URLS:
            url = url_tpl.format(year=year, month=month)
            try:
                req = urllib.request.Request(url, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
                    'Referer': 'https://www.cbr.ru/',
                })
                with urllib.request.urlopen(req, timeout=15) as r:
                    data = r.read()
                if len(data) > 5000:  # валидный Excel
                    print(f"    ✓ ЦБ ПФ Excel: {url[-30:]} ({len(data)//1024} КБ)")
                    return parse_cbr_pf_excel(data)
            except Exception as e:
                pass
    
    print("    → ЦБ РФ недоступен, используем статичные данные")
    return {}

def parse_cbr_pf_excel(data: bytes) -> dict:
    """Парсит Excel ЦБ РФ с данными по банкам ПФ."""
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        
        result = {}
        total = 0
        rows = []
        
        for row in ws.iter_rows(values_only=True):
            if row[0] and isinstance(row[0], str) and any(
                kw in row[0] for kw in ['банк', 'Банк', 'сбер', 'Сбер', 'ВТБ']
            ):
                name = str(row[0]).strip()
                # Пытаемся найти числовые данные
                for val in row[1:]:
                    if isinstance(val, (int, float)) and val > 0:
                        rows.append((name, float(val)))
                        total += float(val)
                        break
        
        if rows and total > 0:
            for name, vol in rows:
                result[name] = {
                    "pf_trn": round(vol / 1e12, 3) if vol > 1e9 else round(vol, 3),
                    "pf_share": round(vol / total * 100, 1),
                }
            print(f"    ✓ Распарсено {len(result)} банков из Excel ЦБ")
        
        return result
    except Exception as e:
        print(f"    WARN parse_cbr_pf_excel: {e}")
        return {}


# ── CLOUDFLARE WORKER ПРОКСИ ─────────────────────────────────────────
# Worker проксирует запросы через российские IP Cloudflare
# Деплой: workers.cloudflare.com (бесплатно, 100k запросов/день)
# После деплоя добавь URL в GitHub Secrets как CBR_PROXY_URL
import os

CF_PROXY_URL = os.environ.get('CBR_PROXY_URL', '')  # из GitHub Secrets

def fetch_via_proxy(url: str) -> bytes:
    """Загружает URL через Cloudflare Worker прокси."""
    if not CF_PROXY_URL:
        return fetch_direct(url)
    
    proxy_url = f"{CF_PROXY_URL.rstrip('/')}/?url={urllib.parse.quote(url, safe='')}"
    try:
        req = urllib.request.Request(proxy_url, headers={
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36',
        })
        with urllib.request.urlopen(req, timeout=20) as r:
            return r.read()
    except Exception as e:
        print(f"    WARN proxy: {e}, пробуем напрямую")
        return fetch_direct(url)

def fetch_direct(url: str) -> bytes:
    """Прямой запрос без прокси."""
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept-Language': 'ru-RU,ru;q=0.9',
            'Referer': 'https://www.cbr.ru/',
        })
        with urllib.request.urlopen(req, timeout=20) as r:
            return r.read()
    except Exception as e:
        print(f"    ERR direct: {e}")
        return b''

def fetch_text_via_proxy(url: str) -> str:
    """Загружает URL через прокси и возвращает текст."""
    data = fetch_via_proxy(url)
    return data.decode('utf-8', errors='replace') if data else ''

# ── ПАРСИНГ ЦБ РФ — ДАННЫЕ ПФ ────────────────────────────────────────
CBR_PF_PAGE = "https://www.cbr.ru/statistics/bank_sector/equity_const_financing/"

def fetch_cbr_pf_live() -> dict:
    """
    Скачивает актуальные данные по ПФ с сайта ЦБ РФ через прокси.
    Возвращает {bank_name: {pf_share, pf_trn, asset_rank}} или {}.
    """
    if not CF_PROXY_URL:
        print("    → CBR_PROXY_URL не задан, используем статику")
        return {}

    print(f"    → Запрос к ЦБ РФ через прокси...")
    
    # Шаг 1: получаем страницу со списком файлов
    html = fetch_text_via_proxy(CBR_PF_PAGE)
    if not html or len(html) < 1000:
        print(f"    → Страница не получена ({len(html)} байт)")
        return {}

    # Шаг 2: ищем ссылку на актуальный Excel файл
    import re
    xlsx_links = re.findall(
        r'href=[^>]*>.*?</a>', html, re.I
    ) or []
    xlsx_links = [x for x in re.findall(r'href="([^"]+\.xlsx)"', html, re.I) if any(k in x.lower() for k in ["pfbalance","pf_banks","equity"])]
    
    if not xlsx_links:
        # Ищем любой xlsx на странице
        xlsx_links = re.findall(r'href="([^"]+\.xlsx?)"', html, re.I)
    
    print(f"    → Найдено Excel файлов: {len(xlsx_links)}")
    
    for link in xlsx_links[:3]:
        # Дополняем relative URLs
        if link.startswith('/'):
            link = 'https://www.cbr.ru' + link
        elif not link.startswith('http'):
            link = 'https://www.cbr.ru/' + link
        
        print(f"    → Скачиваем: {link[-60:]}")
        data = fetch_via_proxy(link)
        
        if data and len(data) > 5000:
            result = parse_cbr_xlsx(data)
            if result:
                print(f"    ✓ Распарсено {len(result)} банков из ЦБ РФ")
                return result
    
    # Если Excel не нашли — парсим HTML таблицу прямо со страницы
    return parse_cbr_html_table(html)

def parse_cbr_xlsx(data: bytes) -> dict:
    """Парсит Excel файл ЦБ РФ с данными по ПФ банков."""
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        
        result = {}
        total_vol = 0
        rows_data = []
        
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                # Пропускаем заголовки
                if any(h in name.lower() for h in ['банк', 'наименование', 'итого', 'всего']):
                    if 'итого' in name.lower() or 'всего' in name.lower():
                        continue
                    # Это строка с данными банка
                    for val in row[1:]:
                        if isinstance(val, (int, float)) and val > 100:  # млн руб
                            rows_data.append((name, float(val)))
                            total_vol += float(val)
                            break
        
        if rows_data and total_vol > 0:
            for name, vol in rows_data:
                # Нормализуем имя банка для совпадения с нашим списком
                normalized = normalize_bank_name(name)
                result[normalized] = {
                    'pf_trn':   round(vol / 1e6, 3),  # конвертируем млн → трлн
                    'pf_share': round(vol / total_vol * 100, 1),
                }
        
        return result
    except Exception as e:
        print(f"    WARN parse_cbr_xlsx: {e}")
        return {}

def parse_cbr_html_table(html: str) -> dict:
    """Парсит HTML таблицу на странице ЦБ РФ если Excel не доступен."""
    import re
    result = {}
    
    # Ищем таблицу с данными банков
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', html, re.S | re.I)
    
    total_vol = 0
    rows_data = []
    
    for row in rows:
        cells = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.S | re.I)
        cells = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
        
        if len(cells) < 2:
            continue
        
        name = cells[0]
        if len(name) < 3 or any(h in name.lower() for h in ['итого', 'всего', '№']):
            continue
        
        # Пытаемся найти числовое значение
        for cell in cells[1:]:
            num_str = cell.replace(' ', '').replace(',', '.')
            try:
                val = float(num_str)
                if val > 100:  # минимальный объём в млрд руб
                    rows_data.append((name, val))
                    total_vol += val
                    break
            except:
                pass
    
    if rows_data and total_vol > 0:
        for name, vol in rows_data:
            normalized = normalize_bank_name(name)
            result[normalized] = {
                'pf_trn':   round(vol / 1000, 3),  # млрд → трлн
                'pf_share': round(vol / total_vol * 100, 1),
            }
    
    return result

def normalize_bank_name(name: str) -> str:
    """Нормализует название банка из ЦБ к нашему формату."""
    name = name.strip()
    mapping = {
        'сбербанк': 'Сбербанк',
        'втб': 'ВТБ',
        'газпромбанк': 'Газпромбанк',
        'альфа': 'Альфа-Банк',
        'россельхоз': 'Россельхозбанк',
        'дом.рф': 'Банк ДОМ.РФ',
        'домрф': 'Банк ДОМ.РФ',
        'промсвязь': 'Промсвязьбанк',
        'псб': 'Промсвязьбанк',
        'совкомбанк': 'Совкомбанк',
        'мкб': 'МКБ',
        'т-банк': 'Т-Технологии',
        'тинькофф': 'Т-Технологии',
        'бспб': 'БСПБ',
        'санкт-петербург': 'БСПБ',
        'мтс': 'МТС Банк',
        'уралсиб': 'Уралсиб',
        'ак барс': 'Ак Барс',
        'новиком': 'Новикомбанк',
        'райффайзен': 'Райффайзенбанк',
    }
    lower = name.lower()
    for key, val in mapping.items():
        if key in lower:
            return val
    return name


def update_pf_data(banks: list) -> list:
    """
    Обновляет данные ПФ для каждого банка.
    1. Пробует ЦБ РФ через Cloudflare Worker прокси
    2. Fallback — статичные данные CBR_PF_SHARES из последнего отчёта
    """
    print("\n── Данные ПФ (ЦБ РФ) ──")
    
    # Пробуем получить свежие данные через прокси
    live_data = fetch_cbr_pf_live()
    
    for b in banks:
        name = b.get("n", "")
        
        # Берём данные: сначала live, потом статика
        pf_info = live_data.get(name) or CBR_PF_SHARES.get(name, {})
        
        b["pf"]         = 1 if pf_info.get("pf_share", 0) > 0 else 0
        b["pf_share"]   = pf_info.get("pf_share")
        b["pf_trn"]     = pf_info.get("pf_trn")
        b["asset_rank"] = pf_info.get("asset_rank", b.get("r", 99))
    
    # Сортируем по доле ПФ
    banks.sort(key=lambda b: (-(b.get("pf_share") or 0), b.get("asset_rank", 99)))
    
    print(f"    ✓ ПФ-данные обновлены для {sum(1 for b in banks if b.get('pf_share'))} банков")
    return banks


def main():
    print(f"{'='*60}")
    print(f"Банковский Радар — агент обновления PDF")
    print(f"{NOW} | pdfplumber: {'OK' if PDF_OK else 'нет'}")
    print(f"{'='*60}\n")

    # Загружаем предыдущий data.json
    # КЛЮЧ = ИМЯ БАНКА (не ранг!) — данные привязаны к названию, не позиции
    prev_latest = {}  # {bank_name: latest_data}
    valid_names = {b["n"] for b in BANKS}
    if OUTPUT.exists():
        try:
            prev = json.loads(OUTPUT.read_text("utf-8"))
            skipped = 0
            for b in prev.get("banks", []):
                name = b.get("n", "")
                # Принимаем только если имя есть в текущем BANKS
                if name in valid_names and b.get("latest"):
                    prev_latest[name] = b["latest"]  # ключ = имя банка
                else:
                    skipped += 1
            print(f"Предыдущие данные: {prev.get('updated_at','?')} "
                  f"({len(prev_latest)} валидных по имени, {skipped} устаревших)\n")
        except Exception as e:
            print(f"WARN prev_data: {e}\n")

    # ── Smart-lab: парсинг LTM данных для биржевых банков ──────────────
    print("\n── Smart-lab LTM ──")
    sl_updates = {}  # {bank_name: data} — ключ = имя банка
    for rank, ticker in SMARTLAB_TICKERS.items():
        # Находим имя банка по рангу
        bank_sl = next((b for b in BANKS if b["r"] == rank), None)
        bank_name_sl = bank_sl["n"] if bank_sl else str(rank)
        print(f"  [{ticker}] {bank_name_sl}...")
        try:
            data = parse_smartlab(ticker)
            if data:
                sl_updates[bank_name_sl] = data  # ключ = имя банка
                print(f"    → {data}")
            else:
                print(f"    → нет данных (возможно блокировка)")
        except Exception as e:
            print(f"    ERR: {e}")
        time.sleep(2)

    # ── Парсинг открытых источников для Ак Барс / Экспобанк / Синара ──
    print("\n── Открытые источники (Ак Барс, Экспобанк, Синара) ──")
    for rank in OPEN_SOURCES:
        bank_name = OPEN_SOURCES[rank]["name"]
        print(f"  {bank_name}...")
        try:
            os_data = parse_open_source_bank(rank)
            if os_data:
                if bank_name not in sl_updates:
                    sl_updates[bank_name] = {}
                sl_updates[bank_name].update(os_data)  # ключ = имя банка
                print(f"    → обновлено: {os_data}")
            else:
                print(f"    → данных не найдено")
        except Exception as e:
            print(f"    ERR {bank_name}: {e}")
        time.sleep(2)

    final_banks = []

    for bank_def in BANKS:
        rank = bank_def["r"]
        base = bank_def["base"]
        est  = bank_def["est"]

        # Пробуем найти новые данные из PDF (e-disclosure + ir_urls)
        latest_from_pdf = None
        if bank_def["st"] not in ("exit", "nd"):
            # Сначала пробуем официальные IR-страницы банка
            ir_urls = bank_def.get("ir_urls", [])
            for ir_url in ir_urls[:2]:
                try:
                    html = fetch(ir_url)
                    if html:
                        text = re.sub(r'<[^>]+>', ' ', html)
                        text = re.sub(r'\s+', ' ', text)
                        ir_metrics = extract_metrics(text)
                        # Ищем PDF ссылки
                        base_domain = re.match(r'https?://[^/]+', ir_url)
                        base = base_domain.group(0) if base_domain else ""
                        pdf_links = re.findall(r'href=["\']([^"\']+\.pdf[^"\']*)["\']|href=["\']([^"\']+msfo[^"\']*)["\']i', html, re.I)
                        for groups in pdf_links[:3]:
                            link = next((g for g in groups if g), None)
                            if not link: continue
                            if link.startswith("/"): link = base + link
                            elif not link.startswith("http"): link = base + "/" + link
                            if ".pdf" in link.lower():
                                pdf_bytes = fetch_binary(link)
                                if pdf_bytes and len(pdf_bytes) > 5000:
                                    pdf_text = extract_pdf_text(pdf_bytes)
                                    if pdf_text:
                                        pdf_m = extract_metrics(pdf_text)
                                        if pdf_m:
                                            ir_metrics.update(pdf_m)
                                time.sleep(1)
                        if ir_metrics:
                            print(f"    ✓ IR-страница {ir_url[:50]}: {ir_metrics}")
                            if bank_def["n"] not in sl_updates: sl_updates[bank_def["n"]] = {}
                            sl_updates[bank_def["n"]].update(ir_metrics)
                            sl_updates[bank_def["n"]]["src"] = f"IR {TODAY}"
                        time.sleep(1.5)
                except Exception as e:
                    print(f"    WARN ir_url {ir_url[:50]}: {e}")
            # Затем e-disclosure
        if bank_def.get("e_id") and bank_def["st"] not in ("exit", "nd"):
            latest_from_pdf = process_bank_pdf(bank_def)

        # Smart-lab данные как fallback если PDF не дал результатов
        sl_data = sl_updates.get(bank_def["n"])  # по имени банка

        # Берём предыдущие latest-данные из кэша
        prev_lat = prev_latest.get(bank_def["n"])  # по имени, не по рангу

        # Выбираем лучший latest
        latest = None
        if latest_from_pdf and latest_from_pdf.get("metrics"):
            latest = {
                "period": latest_from_pdf["period"],
                "period_date": latest_from_pdf["period_date"],
                "src": f"PDF {latest_from_pdf['period']}",
                **{k: v for k,v in latest_from_pdf["metrics"].items()},
            }
        elif sl_data:
            # Используем Smart-lab LTM данные
            latest = {
                "src": f"Smart-lab LTM {TODAY}",
                **sl_data,
            }
        elif prev_lat:
            latest = prev_lat  # сохраняем предыдущий latest

        # Считаем iRoRWA для базовых и latest данных
        def calc_rorwa(roe, h20):
            return round(roe * h20 / 100, 2) if roe and h20 else None

        base_rorwa = calc_rorwa(base.get("roe"), base.get("h20"))

        # Определяем что показывать как основные данные
        # Если latest свежее базы — показываем latest как основные
        use_latest_as_main = (
            latest and
            latest.get("metrics") or (latest and any(latest.get(m) for m in ["roe","nim","cir","cor","h20"]))
        )

        # Основные данные для отображения
        if use_latest_as_main and latest:
            main_roe = latest.get("roe") or base.get("roe")
            main_nim = latest.get("nim") or base.get("nim")
            main_cir = latest.get("cir") or base.get("cir")
            main_cor = latest.get("cor") or base.get("cor")
            main_h20 = latest.get("h20") or base.get("h20")
            main_src = latest.get("src", base["src"])
            main_period = latest.get("period", base["period"])
        else:
            main_roe = base.get("roe")
            main_nim = base.get("nim")
            main_cir = base.get("cir")
            main_cor = base.get("cor")
            main_h20 = base.get("h20")
            main_src = base["src"]
            main_period = base["period"]

        main_rorwa = calc_rorwa(main_roe, main_h20)

        record = {
            "r":   rank,
            "n":   bank_def["n"],
            "lo":  bank_def["lo"],
            "lc":  bank_def["lc"],
            "tp":  bank_def["tp"],
            "pub": bank_def["pub"],
            "pf":  bank_def["pf"],
            # Основные данные (самые свежие)
            "roe":   main_roe,
            "nim":   main_nim,
            "cir":   main_cir,
            "cor":   main_cor,
            "h20":   main_h20,
            "rorwa": main_rorwa,
            "src":   main_src,
            "period": main_period,
            # Флаги оценочности
            "re": est["re"], "ne": est["ne"],
            "ce": est["ce"], "oe": est["oe"], "he": est["he"],
            "st": bank_def["st"],
            # Базовые данные (годовые МСФО) — для сравнения
            "base": {
                "roe":   base.get("roe"),
                "nim":   base.get("nim"),
                "cir":   base.get("cir"),
                "cor":   base.get("cor"),
                "h20":   base.get("h20"),
                "rorwa": base_rorwa,
                "src":   base["src"],
                "period": base["period"],
            },
            # Последние найденные данные (квартальные)
            "latest": latest,
        }
        final_banks.append(record)
        time.sleep(0.5)

    # Обновляем данные ПФ из ЦБ РФ
    final_banks = update_pf_data(final_banks)

    kpi = compute_kpi(final_banks)
    signals = build_signals(final_banks)

    result = {
        "updated":    TODAY,
        "updated_at": NOW,
        "kpi": kpi,
        "banks": final_banks,
        "signals": signals,
        "summary": (
            f"Данные обновлены {TODAY}. "
            f"Медиана ROE ТОП-20: {kpi['median_roe']}%. "
            f"Лидер iRoRWA: {kpi['leader_rorwa_bank']} ({kpi['leader_rorwa_val']}%). "
            f"Мин. буфер капитала: {kpi['min_h20_bank']} (H20.0={kpi['min_h20_val']}%)."
        ),
        "insights": [
            {"icon":"▲","text":f"Лидер iRoRWA: {kpi['leader_rorwa_bank']} — {kpi['leader_rorwa_val']}%"},
            {"icon":"⚠","text":f"Мин. буфер: {kpi['min_h20_bank']}, H20.0={kpi['min_h20_val']}%"},
            {"icon":"★","text":f"Лучший CIR в ПФ: {kpi['best_cir_pf_bank']} — {kpi['best_cir_pf_val']}%"},
            {"icon":"↓","text":"Т-Технологии: ROE 29.1% лидер сектора, COR 6.5% = розничная специфика"},
        ],
        "forecasts": [
            {"icon":"→","text":"Сбер 2026: ROE ~22%, NIM ~5.9%, COR <1.4%"},
            {"icon":"→","text":"ВТБ 2026: ROE 20%, восстановление NIM"},
            {"icon":"→","text":"ДОМ.РФ 2026: прибыль >104 млрд, активы +15%"},
        ],
    }

    OUTPUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), "utf-8")
    banks_with_data = sum(1 for b in final_banks if b.get("roe"))
    banks_with_latest = sum(1 for b in final_banks if b.get("latest") and b["latest"])
    print(f"\n{'='*60}")
    print(f"✓ data.json обновлён")
    print(f"  Банков с ROE: {banks_with_data}/20")
    print(f"  Банков с квартальными данными: {banks_with_latest}/20")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()
